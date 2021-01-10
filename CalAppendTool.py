##************************************************************************************************************************* ##
##************************************************************************************************************************* ##
##                                                                                                                          ##
## Last updated : akuma105 , APPLICATION TEAM GWM                                                                           ##
## Organisation : KPIT Technologies LTd.                                                                                    ##
## About: Tool to udpate static caliberations in C source code                                                              ##
## Script Version: 1.5           																							##
##                                                                                        								    ##
## Evolution History:                                                                                                       ##
##               Version: 1.3 release note                                                                                  ##
##                   1. Colour Codes in Buttons added to reflect validity of associated data                                ##
##                   2. Live display of caliberation being updated added                                                    ##
##                   3. Few containment error prompt messages added.                                                        ##
##               Version: 1.4 release note                                                                                  ##
##                   1. Support for updating Array elements added.                                                          ##
##               Version: 1.5 release note                                                                                  ##
##                   1. Comparison Report generator added                                                                   ##
##					2. GUI updated to accomodate new feature and functionality                                            	##
##                                                                                                                          ##
##                                                                                                                          ##
##**************************************************************************************************************************##
##**************************************************************************************************************************##


import tkinter
from tkinter import ttk
from tkinter import *
import time
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
import xlrd
from datetime import date
import os
from datetime import datetime
import openpyxl
#import pandas
#from PIL import ImageTk, Image

CaliberationCodeFilePath = " "
ExcelWriteCallCount = 0
ExcelWriteCallCountDiffRep = 0
ExcelWriteCallCountDiffDetail = 0
CaliberationParameters = {}
CaliberationParameters1 = {}
CaliberationParameters2 = {}
TuneColumns = ['Select Tune']
ParamsheetParseFlag = 0
percentage = 0
RuninProgressFlag=0
CalSheet1Var = 0
CalSheet2Var = 0
DiffCheckdiff = []
DiffCheckSame = []
DiffCheckdetails = []
TotalCaliberation2Compared=0
BaseTypes = {'uint8': 'U8', 'uint16': 'U16', 'uint32': 'U32', 'uint64': 'U64', 'sint8': 'S8', \
                         'sint16': 'S16', 'sint32': 'S32', 'sint64': 'S64', 'boolean': 'Bool', 'U8': 'U8', \
                         'U16': 'U16', 'U32': 'U32', 'U64': 'U64', 'S8': 'S8', 'S16': 'S16', 'S32': 'S32', \
                         'S64': 'S64', 'Bool': 'Bool'}




def ExcelWrite(ParametersName, FileName, OldValue, NewValue, UpdatedOn, UpdatedAt, TuneColumnSelected):
    global ExcelWriteCallCount
    ExcelWriteCallCount += 1
    wbkName = 'Report.xlsx'
    try:
        wbk = openpyxl.load_workbook(wbkName)
    except:
        ErrorPrompt("File Opening Error", "Unable to Open Report.xlsx\n\nEnsure:\n1. File"
                                          " with same name and format exist\n2. Report.xlsx is not already open "
                                          "in OS explorer\n3. Report Template could be corrupted due to abrupt"
                                          " termination of Calappend in last run. Please restore template ")
    wks = wbk.get_sheet_by_name("UpdateLog")
    
    wks.cell(row=ExcelWriteCallCount + 1, column=1).value = ExcelWriteCallCount
    wks.cell(row=ExcelWriteCallCount + 1, column=2).value = ParametersName
    wks.cell(row=ExcelWriteCallCount + 1, column=3).value = FileName
    wks.cell(row=ExcelWriteCallCount + 1, column=4).value = str(OldValue)
    wks.cell(row=ExcelWriteCallCount + 1, column=5).value = str(NewValue)
    wks.cell(row=ExcelWriteCallCount + 1, column=6).value = UpdatedOn
    wks.cell(row=ExcelWriteCallCount + 1, column=7).value = str(UpdatedAt + " hrs")
    wks.cell(row=ExcelWriteCallCount + 1, column=8).value = TuneColumnSelected

    try:
        wbk.save(wbkName)
        wbk.close
    except:
        ErrorPrompt("File Closing/Saving Error", "Unable to Edit/Save parameter Log in Report.xlsx\n\nEnsure:\n1. File with same name and format exist\n2. Report.xlsx is not already open in OS explorer")

def ComparisonReportExcelWrite(CALName ,CALValue ,Comment ,Result):
    global ExcelWriteCallCountDiffRep
    ExcelWriteCallCountDiffRep += 1
    wbkName = 'Report.xlsx'
    #time.sleep(1)
    try:
        wbk = openpyxl.load_workbook(wbkName)
    except:

       # ErrorPrompt("File Opening Error", "Unable to Open Report.xlsx\n\nEnsure:\n1. File with same name and format exist\n2. Report.xlsx is not already open in OS explorer")
        ErrorPrompt("File Opening Error", "Unable to Open Report.xlsx\n\nEnsure:\n1. File"
                                          " with same name and format exist\n2. Report.xlsx is not already open "
                                          "in OS explorer\n3. Report Template could be corrupted due to abrupt"
                                          " termination of Calappend in last run. Please restore template ")
    wks = wbk.get_sheet_by_name("Comparison Report")
    Datetoday = date.today()
    Timenow = datetime.now()

    wks.cell(row=ExcelWriteCallCountDiffRep + 1, column=1).value = ExcelWriteCallCountDiffRep
    wks.cell(row=ExcelWriteCallCountDiffRep + 1, column=2).value = CALName
    wks.cell(row=ExcelWriteCallCountDiffRep + 1, column=3).value = CALValue
    wks.cell(row=ExcelWriteCallCountDiffRep + 1, column=4).value = Comment
    wks.cell(row=ExcelWriteCallCountDiffRep + 1, column=5).value = str(Datetoday.strftime("%b-%d-%Y"))\
                                                                      +", "+ str(Timenow.strftime("%H:%M:%S"))
    wks.cell(row=ExcelWriteCallCountDiffRep + 1, column=6).value = Result
    try:
        wbk.save(wbkName)
        wbk.close
    except:
        ErrorPrompt("File Closing/Saving Error", "Unable to Edit/Save parameter Log in Report.xlsx\n\nEnsure:\n1. File with same name and format exist\n2. Report.xlsx is not already open in OS explorer")


def ComparisonReportExcelWriteDetails(ParamSheetName1,ParamSheetVersion1,ParamSheetTune1,
                                ParamSheetName2,ParamSheetVersion2,ParamSheetTune2 ):
    global ExcelWriteCallCountDiffDetail
    ExcelWriteCallCountDiffDetail += 1
    wbkName = 'Report.xlsx'
    #time.sleep(1)
    try:
        wbk = openpyxl.load_workbook(wbkName)
    except:
        #ErrorPrompt("File Opening Error", "MSGID124 Unable to Open Report.xlsx\n\nEnsure:\n1. File with same name and format exist\n2. Report.xlsx is not already open in OS explorer")
        ErrorPrompt("File Opening Error", "Unable to Open Report.xlsx\n\nEnsure:\n1. File"
                                          " with same name and format exist\n2. Report.xlsx is not already open "
                                          "in OS explorer\n3. Report Template could be corrupted due to abrupt"
                                          " termination of Calappend in last run. Please restore template ")
    wks = wbk.get_sheet_by_name("Comparison details")
    Datetoday = date.today()
    Timenow = datetime.now()

    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=1).value = ExcelWriteCallCountDiffDetail
    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=2).value = ParamSheetName1
    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=3).value = ParamSheetVersion1
    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=4).value = ParamSheetTune1
    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=5).value = ParamSheetName2
    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=6).value = ParamSheetVersion2
    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=7).value = ParamSheetTune2
    wks.cell(row=ExcelWriteCallCountDiffDetail + 1, column=8).value = str(Datetoday.strftime("%b-%d-%Y"))\
                                                                      +", "+ str(Timenow.strftime("%H:%M:%S"))

    try:
        wbk.save(wbkName)
        wbk.close
    except:
        ErrorPrompt("File Closing/Saving Error", "Unable to Edit/Save parameter Log in Report.xlsx\n\nEnsure:\n1. File with same name and format exist\n2. Report.xlsx is not already open in OS explorer")



def MessageDisplay(Title, Message):
    messagebox.showinfo(Title, Message)

def SheetButtonConfig(Colour):
    ParamSheetButton.configure(bg = Colour)
    GUITopFrame.update()
    
def ParseButtonConfig(Colour):
    ParseButton.configure(bg = Colour)
    GUITopFrame.update()
    
def CodeButtonConfig(Colour):
    CodeButton.configure(bg = Colour)
    GUITopFrame.update()


def BrowseCalSheet1ButtonConfig(Colour):
    BrowseCalSheet1.configure(bg=Colour)
    GUITopFrame.update()

def BrowseCalSheet2ButtonConfig(Colour):
    BrowseCalSheet2.configure(bg = Colour)
    GUITopFrame.update()


def RunButtonConfig(Colour):
    RunButtonUpdate.configure(bg = Colour)
    GUITopFrame.update()

def GenerateButtonConfig(Colour):
    GenerateDiffReport.configure(bg = Colour)
    GUITopFrame.update()


def UpdateCurrentCAL(CALcurr):
    CurrentCALUpdate.configure(text= str(CALcurr))
    GUITopFrame.update()

def HandleCodefile():
    global RuninProgressFlag,arrayInProcess
    arrayInProcess=0
    #RuninProgressFlag
    # print(ParamsheetParseFlag)
    if ParamsheetParseFlag != 1:
        ParseButtonConfig("red")
        ErrorPrompt("Step Missed", "Parse the Parameter Sheet before initiating Run")
        return
    elif "." not in CaliberationCodeFilePath:
        CodeButtonConfig("red")
        ErrorPrompt("Step Missed", "Code File Not Selected")

        return
    FileName = str(str(CaliberationCodeFilePath).split("/")[-1])
    InputFilepath = CaliberationCodeFilePath
    Outputfileobject = []
    if ".c" in FileName:
        global FileCaliberations, CalsectionstartFlag
        FileCaliberations = []
        CalsectionstartFlag = 0
        if os.path.isdir("Output") == True:
            pass
        else:
            os.makedirs('Output/inc')
            os.makedirs('Output/src')
        CurrentDirectory = str(os.getcwd())
        # print(CurrentDirectory)
        Datetoday = date.today()
        Timenow = datetime.now()
        Inputfile = open(InputFilepath, 'r')
        OutputFileName = FileName

        if ".c" in OutputFileName:
            Outputfilepath = str(CurrentDirectory + "\\" + "Output" + "\\" + "src" + "\\" + OutputFileName)
        elif ".h" in OutputFileName:
            Outputfilepath = str(CurrentDirectory + "\\" + "Output" + "\\" + "inc" + "\\" + OutputFileName)
        else:
            # Display Error Msg
            pass
        Outputfile = open(Outputfilepath, 'w')
        Inputfileobject = Inputfile.readlines()

        TotalCaliberationsCount = 0
        CaliberationUpdatedCount = 0

        for line in Inputfileobject:
            if ";" in line:
                TotalCaliberationsCount += 1
            else:
                pass

        UpdateProgressStatus(TotalCaliberationsCount, 0)
        RunButtonConfig("green")
        for line in Inputfileobject:
            if "Generation date" in line:
                TempList = line.split(":")
                Outputfileobject.append(str(TempList[0] + " :" + str(Datetoday.strftime("%b-%d-%Y")) + \
                                            "  " + str(Timenow.strftime("%H:%M:%S")) + "\n"))
            elif "#pragma" and "const" in line:
                Outputfileobject.append("#pragma ghs section rodata = \".FLASH_CAL_DATA\"" + "\n")
                CalsectionstartFlag = 1;
            elif "__declspec" in line:
                pass

            elif "CAL " and "=" in line and "[" not in line:
                RuninProgressFlag = 1
                # Outputfileobject.append(line)
                # pass
                # print(line)
                LocalLineList = str(line).split()
                LocalLineList[1] = str(BaseTypes[LocalLineList[1]])
                CurrCAL = LocalLineList[2]
                if str(LocalLineList[2]) in CaliberationParameters.keys():
                    # print(LocalLineList)
                    ExcelWrite(str(LocalLineList[2]), FileName, str(LocalLineList[4]), \
                               (str(CaliberationParameters[str(LocalLineList[2])]).partition(".")[0])+" Updated" \
                               , str(Datetoday.strftime("%b-%d-%Y")), \
                               str(Timenow.strftime("%H:%M:%S")), str(SelectedCaliberationTune))

                    LocalLineList[4] = str(CaliberationParameters[str(LocalLineList[2])]).partition(".")[0]
                    # print(LocalLineList)
                    FileCaliberations.append(LocalLineList[1] + \
                                             " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                 4] + ";\n")
                    LocalLineList[2] = "ROM0_" + str(LocalLineList[2])

                    Outputfileobject.append(LocalLineList[0] + " " + LocalLineList[1] + \
                                            " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                4] + ";\n")
                else:
                    #print(LocalLineList[2])
                    ExcelWrite(str(LocalLineList[2]), FileName, str(LocalLineList[4]), \
                               str(LocalLineList[4]+" Retained"), str(Datetoday.strftime("%b-%d-%Y")), \
                               str(Timenow.strftime("%H:%M:%S")), str(SelectedCaliberationTune))

                    FileCaliberations.append(LocalLineList[1] + \
                                             " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                 4] + ";\n")
                    LocalLineList[2] = "ROM0_" + str(LocalLineList[2])
                    Outputfileobject.append(LocalLineList[0] + " " + LocalLineList[1] + \
                                            " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                4] + ";\n")

                CaliberationUpdatedCount += 1
                UpdateProgressStatus(TotalCaliberationsCount, CaliberationUpdatedCount)
                UpdateCurrentCAL(CurrCAL)

            elif "[" and "CAL " and "=" in line:
                arrayInProcess = 1
                LocalLineList = str(line).split()
                LocalLineList[1] = str(BaseTypes[LocalLineList[1]])
                CurrCAL = LocalLineList[2]
                LocalLineList.append("")

                if str(str(LocalLineList[2]).partition("[")[0]) in CaliberationParameters.keys():
                    LocalLineList[4] = str(CaliberationParameters[str(str(LocalLineList[2]).partition("[")[0])])
                    #print(LocalLineList[4])
                    if "," in str(LocalLineList[4]):
                        arrayprocessed = str(LocalLineList[4]).split(",")
                    else:
                        arrayprocessed = str(LocalLineList[4]).split(" ")
                    #print(arrayprocessed)
                    arrayprocessed[0] = "{ "
                    arrayprocessed[-1] = " }"
                    #print(arrayprocessed)
                    LocalLineList[4]=""
                    i=0
                    for  i in range(len(arrayprocessed)-1):
                        if i==0:
                            LocalLineList[4] = "\n"+str(arrayprocessed[i])
                        elif i == len(arrayprocessed)-2:
                            LocalLineList[4] += str(arrayprocessed[i]) + str(arrayprocessed[i+1])
                        else :
                            if (i%17)==0:
                                LocalLineList[4] += str(arrayprocessed[i]) + ", \n"
                            else:
                                LocalLineList[4] += str(arrayprocessed[i]) + ", "
                    ExcelWrite(str(LocalLineList[2]), FileName, "TLDR", \
                               LocalLineList[4] + " Updated" \
                               , str(Datetoday.strftime("%b-%d-%Y")), \
                               str(Timenow.strftime("%H:%M:%S")), str(SelectedCaliberationTune))


                    FileCaliberations.append(LocalLineList[1] + \
                                             " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                 4] + ";\n")
                    LocalLineList[2] = "ROM0_" + str(LocalLineList[2])

                    Outputfileobject.append(LocalLineList[0] + " " + LocalLineList[1] + \
                                            " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                4] + ";\n")
                else:
                    arrayInProcess = 0
                   # print(LocalLineList[2])
                    ExcelWrite(str(LocalLineList[2]), FileName, "TLDR", \
                              str("TLDR" + " Retained"), str(Datetoday.strftime("%b-%d-%Y")), \
                              str(Timenow.strftime("%H:%M:%S")), str(SelectedCaliberationTune))

                    FileCaliberations.append(LocalLineList[1] + \
                                             " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                 4] + ";\n")
                    LocalLineList[2] = "ROM0_" + str(LocalLineList[2])
                    Outputfileobject.append(LocalLineList[0] + " " + LocalLineList[1] + \
                                            " " + LocalLineList[2] + " " + LocalLineList[3] + " " + LocalLineList[
                                                4] + ";\n")

                CaliberationUpdatedCount += 1
                UpdateProgressStatus(TotalCaliberationsCount, CaliberationUpdatedCount)
                UpdateCurrentCAL(CurrCAL)
            elif arrayInProcess == 1 and ";" in line:
                arrayInProcess = 0
            elif arrayInProcess == 1:
                pass

            elif "#pragma" and ".default" in line:
                Outputfileobject.append("#pragma ghs section rodata = default\n\n")
                Outputfileobject.append("#pragma ghs section data = \".LLSW_RAM_CALIBRATION\"\n")
                for element in FileCaliberations:
                    Outputfileobject.append(element)
                Outputfileobject.append("#pragma ghs section data = default\n")
                CalsectionstartFlag = 0

            else:
                Outputfileobject.append(line)
                if CalsectionstartFlag != 0:
                    FileCaliberations.append(line)

        for element in Outputfileobject:
            Outputfile.write(element)
        Outputfile.close()
        UpdateCurrentCAL(str("  Caliberations Generated !!!"))
        MessageDisplay("Caliberation Update Status", str("Successful\n\nOutput Directory : \n" + Outputfilepath))
    elif ".h" in FileName:
        if os.path.isdir("Output") == True:
            pass
        else:
            os.makedirs('Output/inc')
            os.makedirs('Output/src')
        CurrentDirectory = str(os.getcwd())
        # print(CurrentDirectory)
        Datetoday = date.today()
        Timenow = datetime.now()
        Inputfile = open(InputFilepath, 'r')
        OutputFileName = FileName

        if ".c" in OutputFileName:
            Outputfilepath = str(CurrentDirectory + "\\" + "Output" + "\\" + "src" + "\\" + OutputFileName)
        elif ".h" in OutputFileName:
            Outputfilepath = str(CurrentDirectory + "\\" + "Output" + "\\" + "inc" + "\\" + OutputFileName)
        else:
            # Display Error Msg
            pass
        try:
            Outputfile = open(Outputfilepath, 'w')
        except:
            ErrorPrompt("File Open Error","Unable to Open /Access Output file")
        Inputfileobject = Inputfile.readlines()
        TotalCaliberationsCount = 0
        CaliberationUpdatedCount = 0
        for line in Inputfileobject:
            if ";" in line:
                TotalCaliberationsCount += 1
            else:
                pass
        UpdateProgressStatus(TotalCaliberationsCount, 0)
        RunButtonConfig("green")
        for line in Inputfileobject:
            if "Generation date" in line:
                TempList = line.split(":")
                Outputfileobject.append(str(TempList[0] + " :" + str(Datetoday.strftime("%b-%d-%Y")) + \
                                            "  " + str(Timenow.strftime("%H:%M:%S")) + "\n"))
            elif "extern CAL " in line:
                RuninProgressFlag=1
                LocalLineList = str(line).split()
                del LocalLineList[1]
                Outputfileobject.append(LocalLineList[0] + " " + LocalLineList[1] + " " + LocalLineList[2] + ";\n")
                CaliberationUpdatedCount += 1
                UpdateProgressStatus(TotalCaliberationsCount, CaliberationUpdatedCount)
                UpdateCurrentCAL(str(LocalLineList[2]))
            else:
                Outputfileobject.append(line)

        for element in Outputfileobject:
            Outputfile.write(element)
        Outputfile.close()
        UpdateCurrentCAL(str("  Caliberations Generated !!!"))
        MessageDisplay("Caliberation Update Status", str("Successful\n\nOutput Directory : \n" + Outputfilepath))

    else:
        ErrorPrompt("Input Error", "Wrong Model generated code file selected (.h/.c)")
        CodeButtonConfig("red")
    RuninProgressFlag = 0

def ParameterSheetSelect():
    UpdateCurrentCAL(str(" "))
    ParseButtonConfig("steel blue")
    CodeButtonConfig("steel blue")
    RunButtonConfig("steel blue")
    global ParameterSheetPath, EndOfDataRow, EndOfDataCol, CaliberationTuneName, TuneSelected,TuneColmCol
    UpdateProgressStatus(0, 0)
    ParameterSheetPath = str(askopenfilename(title='Select Caliberation Parameter Sheet '))
    if ".xl" and "param" not in ParameterSheetPath:
        SheetButtonConfig("red")
        ErrorPrompt("Input Error",
                    "File selcted with incorrect file format\nExpected file format: XXXX_params.xls/xlsm")
        return

    else:
        try:
            Workbook = xlrd.open_workbook(ParameterSheetPath)  # xls file to read from
            WorkbookSheet = Workbook.sheet_by_index(0)  # first sheet in workbook
            for row in range(WorkbookSheet.nrows):
                for col in range(WorkbookSheet.ncols):
                    global TuneColmCol, TuneColmRow
                    if WorkbookSheet.cell_value(row, col) == 'Tune Name':
                        TuneColmCol = col + 1
                        TuneColmRow = row
            for row in range(WorkbookSheet.nrows):
                for col in range(WorkbookSheet.ncols):
                    if WorkbookSheet.cell_value(row, col) == 'End Of Data':
                        EndOfDataRow = row
                        EndOfDataCol = col
        except:
            ErrorPrompt("Input Error",
                        "File seelcted with incorrect file format\nExpected file format: XXXX_params.xls/xlsm")
            SheetButtonConfig("red")
        TuneColumns.clear()

        while TuneColmCol < EndOfDataCol:
            TuneColumns.append(str(WorkbookSheet.cell_value(TuneColmRow, TuneColmCol)))
            TuneColmCol += 1

        TuneSelected = ttk.Combobox(labelframe1, cursor='arrow', values=TuneColumns, width=25, state="enabled",
                                    justify="left")
        TuneSelected.grid(column=2, row=0)
        try:
            TuneSelected.current(0)
        except:
            TuneSelected.current("Select Tune")

        TuneSelected.bind("<<ComboboxSelected>>",UpdateTuneSelected)
        UpdateTuneSelected(" ")
        SheetButtonConfig("green")

def DiferenceParamsheetRead(SheetNum):
    global ParameterSheetPath1,ParameterSheetPath2, EndOfDataRow, EndOfDataCol,\
        CaliberationTuneName, TuneSelected1,TuneSelected2, TuneColmCol
    global TuneColmCol, TuneColmRow
    UpdateProgressStatus(100, 0)
    UpdateCurrentCAL(" ")
    if SheetNum==1:
        GenerateButtonConfig("Steel Blue")
        #UpdateProgressStatus(0, 0)
        ParameterSheetPath1 = str(askopenfilename(title='Select Caliberation Parameter Sheet '))
        if ".xl" and "param" not in ParameterSheetPath1:
            BrowseCalSheet1ButtonConfig("red")
            ErrorPrompt("Input Error",
                    "File selected with incorrect file format\nExpected file format: XXXX_params.xls/xlsm")
            return

        else:
            try:
                Workbook = xlrd.open_workbook(ParameterSheetPath1)  # xls file to read from
                WorkbookSheet = Workbook.sheet_by_index(0)  # first sheet in workbook
                for row in range(WorkbookSheet.nrows):
                    for col in range(WorkbookSheet.ncols):
                        if WorkbookSheet.cell_value(row, col) == 'Tune Name':
                            TuneColmCol = col + 1
                            TuneColmRow = row
                for row in range(WorkbookSheet.nrows):
                    for col in range(WorkbookSheet.ncols):
                        if WorkbookSheet.cell_value(row, col) == 'End Of Data':
                            EndOfDataRow = row
                            EndOfDataCol = col
            except:
                ErrorPrompt("Input Error",
                        "File seelcted with incorrect file format\nExpected file format: XXXX_params.xls/xlsm")
                BrowseCalSheet1ButtonConfigSheetButtonConfig("red")
            TuneColumns.clear()

            while TuneColmCol < EndOfDataCol:
                TuneColumns.append(str(WorkbookSheet.cell_value(TuneColmRow, TuneColmCol)))
                TuneColmCol += 1

            TuneSelected1 = ttk.Combobox(labelframe2, cursor='arrow', values=TuneColumns, width=26, state="enabled",
                                    justify="left")
            TuneSelected1.grid(column=3, row=0)
            try:
                TuneSelected1.current(0)
            except:
                TuneSelected1.current("Select Tune")

            TuneSelected1.bind("<<ComboboxSelected>>", UpdateTuneSelected1)
            UpdateTuneSelected1(" ")
            BrowseCalSheet1ButtonConfig("green")
    else :
        GenerateButtonConfig("Steel Blue")
        #UpdateProgressStatus(0, 0)
        ParameterSheetPath2 = str(askopenfilename(title='Select Caliberation Parameter Sheet '))
        if ".xl" and "param" not in ParameterSheetPath2:
            BrowseCalSheet2ButtonConfig("red")
            ErrorPrompt("Input Error",
                    "File selected with incorrect file format\nExpected file format: XXXX_params.xls/xlsm")
            return

        else:
            try:
                Workbook = xlrd.open_workbook(ParameterSheetPath2)  # xls file to read from
                WorkbookSheet = Workbook.sheet_by_index(0)  # first sheet in workbook
                for row in range(WorkbookSheet.nrows):
                    for col in range(WorkbookSheet.ncols):

                        if WorkbookSheet.cell_value(row, col) == 'Tune Name':
                            TuneColmCol = col + 1
                            TuneColmRow = row
                for row in range(WorkbookSheet.nrows):
                    for col in range(WorkbookSheet.ncols):
                        if WorkbookSheet.cell_value(row, col) == 'End Of Data':
                            EndOfDataRow = row
                            EndOfDataCol = col
            except:
                ErrorPrompt("Input Error",
                        "File seelcted with incorrect file format\nExpected file format: XXXX_params.xls/xlsm")
                BrowseCalSheet2ButtonConfig("red")
            TuneColumns.clear()

            while TuneColmCol < EndOfDataCol:
                TuneColumns.append(str(WorkbookSheet.cell_value(TuneColmRow, TuneColmCol)))
                TuneColmCol += 1

            TuneSelected2 = ttk.Combobox(labelframe2, cursor='arrow', values=TuneColumns, width=26, state="enabled",
                                    justify="left")
            TuneSelected2.grid(column=3, row=1)
            try:
                TuneSelected2.current(0)
            except:
                TuneSelected2.current("Select Tune")

            TuneSelected2.bind("<<ComboboxSelected>>", UpdateTuneSelected2)
            UpdateTuneSelected2(" ")
            BrowseCalSheet2ButtonConfig("green")

def ErrorPrompt(ErrorType, message):
    messagebox.showwarning(ErrorType, message)


def UpdateCaliberationsCodeFiles():
    UpdateProgressStatus(0, 0)
    try:
        global CaliberationCodeFilePath
        UpdateCurrentCAL(str(" "))
        CodeButtonConfig("green")
        RunButtonConfig("steel blue")
        CaliberationCodeFilePath = str(askopenfilename(title='Select Model generated Caliberation Code file : '))
        # HandleCodefile(str(os.path.join(InputFileslocation, file)))
    except:
        pass


def UpdateProgressStatus(MaxVal, Currprogress):
    global percentage
    ProgressStatus["maximum"] = MaxVal
    ProgressStatus["value"] = Currprogress
    try:
        percentage = round(Currprogress / MaxVal * 100)
        # print(percentage)
    except:
        percentage = 0
    style.configure('text.Horizontal.TProgressbar', text="Progress: " + str(percentage) + '%')
    GUITopFrame.update()


def UpdateTuneSelected(event):
    global SelectedCaliberationTune, ParamsheetParseFlag
    if RuninProgressFlag !=1:
        SelectedCaliberationTune = TuneSelected.get()
        #print("Tune Selected: " + SelectedCaliberationTune)
        UpdateProgressStatus(0, 0)
        ParamsheetParseFlag = 0
        ParseButtonConfig("steel blue")
        CodeButtonConfig("steel blue")
        RunButtonConfig("steel blue")
        UpdateCurrentCAL(str(" "))
        # print(ParamsheetParseFlag)
    else:
        ErrorPrompt("Invalid Operation" , "Caliberation update in Progress..\nTune unmodifiable for current Runtime. ")

def UpdateTuneSelected1(event):
    global SelectedCaliberationTune1, ParamsheetParseFlag
    if RuninProgressFlag !=1:
        SelectedCaliberationTune1 = TuneSelected1.get()
        #print("Tune SelectedUpdateTuneSelected1: " + SelectedCaliberationTune1)
        UpdateProgressStatus(0, 0)
        ParamsheetParseFlag = 0
        #ParseButtonConfig("steel blue")
        #CodeButtonConfig("steel blue")
        GenerateButtonConfig("steel blue")
        UpdateCurrentCAL(str(" "))
        # print(ParamsheetParseFlag)
    else:
        ErrorPrompt("Invalid Operation" , "Caliberation update in Progress..\nTune unmodifiable for current Runtime. ")




#*****************************************************************************************************************
def GenerateComparisonReport ():
    global TotalCaliberation2Compared
    UpdateCurrentCAL("Initiating Comparison ... ")
    ParsingSheet = 1
    if RuninProgressFlag == 1:
        ErrorPrompt("Invalid Operation", "Caliberation update in Progress..\n Re-Parsing unavailable.  ")
        return
    else:
        pass

    #UpdateCurrentCAL(str(" "))
    global TuneCol, ParameterListCol, ParameterListRow, CalculatedVal, CaliberationParameters1, ParamsheetParseFlag
    CalculatedVal = 0.0
    Workbook = xlrd.open_workbook(ParameterSheetPath1)  # xls file to read from
    WorkbookSheet = Workbook.sheet_by_index(0)  # first sheet in workbook

    for row in range(WorkbookSheet.nrows):
        for col in range(WorkbookSheet.ncols):
            if WorkbookSheet.cell_value(row, col) == 'Parameter Name':
                ParameterListCol = col
                ParameterListRow = row

    for row in range(WorkbookSheet.nrows):
        for col in range(WorkbookSheet.ncols):
            if WorkbookSheet.cell_value(row, col) == str(SelectedCaliberationTune1):
                TuneCol = col
    #print(SelectedCaliberationTune1)
    while ParameterListRow < EndOfDataRow:
        ParameterListRow += 1
        try:
            CaliberationParameters1[str(WorkbookSheet.cell_value(ParameterListRow, ParameterListCol))] = str(
                WorkbookSheet.cell_value(ParameterListRow, TuneCol));
        except:
            #ParseButtonConfig("red")
            ErrorPrompt("Parsing Error", "MSGID591: Error in Parsing and fetching Parameter values in selecetd Parameter Sheet")

    for Key in CaliberationParameters1:
        if str(CaliberationParameters1[Key]).find("+") != -1:
            TempElementList = str(CaliberationParameters1[Key]).split("+")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters1[str(element)]

            for index in TempElementList:
                CalculatedVal += float(index)
            CaliberationParameters1[Key] = CalculatedVal

    CalculatedVal = 0.0

    for Key in CaliberationParameters1:

        if str(CaliberationParameters1[Key]).find("-") != -1:
            TempElementList = str(CaliberationParameters1[Key]).split("+")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters1[str(element)]

            for index in TempElementList:
                CalculatedVal -= float(index)
            CaliberationParameters1[Key] = CalculatedVal

    CalculatedVal = 1.0

    for Key in CaliberationParameters1:
        if str(CaliberationParameters1[Key]).find("*") != -1:

            TempElementList = str(CaliberationParameters1[Key]).split("*")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    # print(TempElementList.index(element))
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters1[str(element)]

            for element in TempElementList:
                CalculatedVal *= float(element)
            CaliberationParameters1[Key] = CalculatedVal

    CalculatedVal = 0.0

    for Key in CaliberationParameters1:
        if str(CaliberationParameters1[Key]).find("/") != -1:
            TempElementList = str(CaliberationParameters1[Key]).split("/")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    # print(TempElementList.index(element))
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters1[str(element)]
            indexDiv = 0
            while indexDiv in range(len(TempElementList)) and int(indexDiv) < len(TempElementList) - 1:
                CalculatedVal = (float(TempElementList[indexDiv]) / float(TempElementList[indexDiv + 1]))
                indexDiv += 1
            CaliberationParameters1[Key] = CalculatedVal
    CalculatedVal = 0.0

    # print(CaliberationParameters1)

    for Key in CaliberationParameters1:
        if re.search('[a-zA-Z]', str(CaliberationParameters1[Key])) is not None:
            # print(re.search('[a-zA-Z]', str(CaliberationParameters1[Key])))
            try:
                CaliberationParameters1[Key] = CaliberationParameters1[str(CaliberationParameters1[Key])];
            except:
                #ParseButtonConfig("red")
                # print(Key)
                ErrorPrompt("Parsing Error", "MSGID482 Error parsing Parameter values in Provided Parameter Sheet")

    for Key in CaliberationParameters1:
        if str(CaliberationParameters1[Key]).find(",") != -1:
            try:
                CaliberationParameters1[Key] = str(CaliberationParameters1[Key]).replace("[", "{").replace("]", "}");
            except:
                #ParseButtonConfig("red")
                ErrorPrompt("Parsing Error", "MSGID490 Error parsing Parameter values in Provided Parameter Sheet")

    # print(CaliberationParameters1)
    #ParamsheetParseFlag = 1
    #ParseButtonConfig("green")
    #print(CaliberationParameters1)
    CalculatedVal = 0.0
    Workbook = xlrd.open_workbook(ParameterSheetPath2)  # xls file to read from
    WorkbookSheet = Workbook.sheet_by_index(0)  # first sheet in workbook

    for row in range(WorkbookSheet.nrows):
        for col in range(WorkbookSheet.ncols):
            if WorkbookSheet.cell_value(row, col) == 'Parameter Name':
                ParameterListCol = col
                ParameterListRow = row

    for row in range(WorkbookSheet.nrows):
        for col in range(WorkbookSheet.ncols):
            if WorkbookSheet.cell_value(row, col) == str(SelectedCaliberationTune2):
                TuneCol = col
    #print(SelectedCaliberationTune2)
    while ParameterListRow < EndOfDataRow:
        ParameterListRow += 1
        try:
            CaliberationParameters2[str(WorkbookSheet.cell_value(ParameterListRow, ParameterListCol))] = str(
                WorkbookSheet.cell_value(ParameterListRow, TuneCol));
        except:
            #ParseButtonConfig("red")
            ErrorPrompt("Parsing Error", "MSGID592: Error in Parsing and fetching Parameter values in selecetd Parameter Sheet")

    for Key in CaliberationParameters2:
        if str(CaliberationParameters2[Key]).find("+") != -1:
            TempElementList = str(CaliberationParameters2[Key]).split("+")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters2[str(element)]

            for index in TempElementList:
                CalculatedVal += float(index)
            CaliberationParameters2[Key] = CalculatedVal

    CalculatedVal = 0.0

    for Key in CaliberationParameters2:

        if str(CaliberationParameters2[Key]).find("-") != -1:
            TempElementList = str(CaliberationParameters2[Key]).split("+")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters2[str(element)]

            for index in TempElementList:
                CalculatedVal -= float(index)
            CaliberationParameters2[Key] = CalculatedVal

    CalculatedVal = 1.0

    for Key in CaliberationParameters2:
        if str(CaliberationParameters2[Key]).find("*") != -1:

            TempElementList = str(CaliberationParameters2[Key]).split("*")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    # print(TempElementList.index(element))
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters2[str(element)]

            for element in TempElementList:
                CalculatedVal *= float(element)
            CaliberationParameters2[Key] = CalculatedVal

    CalculatedVal = 0.0

    for Key in CaliberationParameters2:
        if str(CaliberationParameters2[Key]).find("/") != -1:
            TempElementList = str(CaliberationParameters2[Key]).split("/")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    # print(TempElementList.index(element))
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters2[str(element)]
            indexDiv = 0
            while indexDiv in range(len(TempElementList)) and int(indexDiv) < len(TempElementList) - 1:
                CalculatedVal = (float(TempElementList[indexDiv]) / float(TempElementList[indexDiv + 1]))
                indexDiv += 1
            CaliberationParameters2[Key] = CalculatedVal
    CalculatedVal = 0.0

    # print(CaliberationParameters2)

    for Key in CaliberationParameters2:
        if re.search('[a-zA-Z]', str(CaliberationParameters2[Key])) is not None:
            # print(re.search('[a-zA-Z]', str(CaliberationParameters2[Key])))
            try:
                CaliberationParameters2[Key] = CaliberationParameters2[str(CaliberationParameters2[Key])];
            except:
                #ParseButtonConfig("red")
                # print(Key)
                ErrorPrompt("Parsing Error", "MSGID775 Error parsing Parameter values in Provided Parameter Sheet")

    for Key in CaliberationParameters2:
        if str(CaliberationParameters2[Key]).find(",") != -1:
            try:
                CaliberationParameters2[Key] = str(CaliberationParameters2[Key]).replace("[", "{").replace("]", "}");
            except:
                #ParseButtonConfig("red")
                ErrorPrompt("Parsing Error", "MSGID783 Error parsing Parameter values in Provided Parameter Sheet")

    # print(CaliberationParameters2)
    #ParamsheetParseFlag = 1
    #ParseButtonConfig("green")
    #print(CaliberationParameters2)

    ComparisonReportExcelWriteDetails(ParameterSheetPath1,EntryCalSheet1Var.get(),SelectedCaliberationTune1,
                                ParameterSheetPath2,EntryCalSheet2Var.get(),SelectedCaliberationTune2)
    TotalCaliberation2CompareCount = len(CaliberationParameters1.keys())+len(CaliberationParameters2.keys())

    for Key in CaliberationParameters1:
        UpdateCurrentCAL("Comparing: " + Key)
        if Key in CaliberationParameters2.keys():
            if CaliberationParameters2[Key] == CaliberationParameters1[Key]:
                #print("Kay Value Same")
               # DiffCheckSame.append([Key, str(CaliberationParameters1[Key]) + " ,  " + str(CaliberationParameters2[Key]),
                                    #  "Caliberation parameter value similar in parameter sheet 1 and 2"])
                ComparisonReportExcelWrite(Key, str(CaliberationParameters1[Key]) + " , " + str(
                    CaliberationParameters2[Key]),
                                           "Caliberation parameter value Similar in parameter sheet 1 and 2","Similar")
            else:
                #DiffCheckdiff.append([Key, str(CaliberationParameters1[Key])+" , "+str(CaliberationParameters2[Key]),
                                    # "Caliberation parameter value differs in parameter sheet 1 and 2"])

               #df = pandas.DataFrame(dataframe, columns=['Caliberation Name', 'Caliberation Value', 'Comment'])
               #print(df)
                ComparisonReportExcelWrite(Key, str(CaliberationParameters1[Key])+" , "+str(CaliberationParameters2[Key]),
                                     "Caliberation parameter value differs in parameter sheet 1 and 2","Different")
        else:
            #print(Key)
            #DiffCheckdiff.append([Key, str(CaliberationParameters1[Key]),
                                  #"Caliberation parameter present in parameter sheet 1 but absent in parameter sheet 2"])
            ComparisonReportExcelWrite(str(Key) ,str(CaliberationParameters1[Key]),
                                 "Caliberation parameter present in parameter sheet 1 but absent in parameter sheet 2" ,"Different")
        TotalCaliberation2Compared += 1
        UpdateProgressStatus(TotalCaliberation2CompareCount,TotalCaliberation2Compared)
    for Key in CaliberationParameters2:
        UpdateCurrentCAL("Comparing: " + Key)
        if Key in CaliberationParameters1.keys():
            pass
        else:
            #DiffCheckdiff.append([Key, str(CaliberationParameters2[Key]),
                             #     "Caliberation parameter present in parameter sheet 2 but absent in parameter sheet 1"])
            ComparisonReportExcelWrite(str(Key), str(CaliberationParameters2[Key]),
                                 "Caliberation parameter present in parameter sheet 2 but absent in parameter sheet 1","Different")
        TotalCaliberation2Compared += 1
        UpdateProgressStatus(TotalCaliberation2CompareCount,TotalCaliberation2Compared)
        GenerateButtonConfig("green")
    UpdateCurrentCAL("Caliberations Comparison Successful.")
    MessageDisplay("Caliberation Comparsion Status", str("Successful\nReport updated "))
    #DataFramedif= pandas.DataFrame(DiffCheckdiff)
    #DataFrameSame= pandas.DataFrame(DiffCheckSame)

    #print(DataFramedif)
    #print(DataFrameSame)
#*****************************************************************************************************************





def UpdateTuneSelected2(event):
    global SelectedCaliberationTune2, ParamsheetParseFlag
    if RuninProgressFlag !=1:
        SelectedCaliberationTune2 = TuneSelected2.get()
        # print("Tune Selected: " + SelectedCaliberationTune)
        UpdateProgressStatus(0, 0)
        ParamsheetParseFlag = 0
        #ParseButtonConfig("steel blue")
        #CodeButtonConfig("steel blue")
        GenerateButtonConfig("steel blue")
        UpdateCurrentCAL(str(" "))
        # print(ParamsheetParseFlag)
    else:
        ErrorPrompt("Invalid Operation" , "Caliberation update in Progress..\nTune unmodifiable for current Runtime. ")

def ParseNCalculateCaliberations():
    if RuninProgressFlag == 1:
        ErrorPrompt("Invalid Operation", "Caliberation update in Progress..\n Re-Parsing unavailable.  ")
        return
    else:
        pass

    UpdateCurrentCAL(str(" "))
    global TuneCol, ParameterListCol, ParameterListRow, CalculatedVal, CaliberationParameters, ParamsheetParseFlag
    CalculatedVal = 0.0
    Workbook = xlrd.open_workbook(ParameterSheetPath)  # xls file to read from
    WorkbookSheet = Workbook.sheet_by_index(0)  # first sheet in workbook

    for row in range(WorkbookSheet.nrows):
        for col in range(WorkbookSheet.ncols):
            if WorkbookSheet.cell_value(row, col) == 'Parameter Name':
                ParameterListCol = col
                ParameterListRow = row

    for row in range(WorkbookSheet.nrows):
        for col in range(WorkbookSheet.ncols):
            if WorkbookSheet.cell_value(row, col) == str(SelectedCaliberationTune):
                TuneCol = col
                    
    while ParameterListRow < EndOfDataRow:
        ParameterListRow += 1
        try:
            CaliberationParameters[str(WorkbookSheet.cell_value(ParameterListRow, ParameterListCol))] = str(WorkbookSheet.cell_value(ParameterListRow, TuneCol));
        except:
            ParseButtonConfig("red")
            ErrorPrompt("MSGID728: Parsing Error", "Error in Parsing and fetching Parameter values in selecetd Parameter Sheet")

    for Key in CaliberationParameters:
        if str(CaliberationParameters[Key]).find("+") != -1:
            TempElementList = str(CaliberationParameters[Key]).split("+")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters[str(element)]

            for index in TempElementList:
                CalculatedVal += float(index)
            CaliberationParameters[Key] = CalculatedVal

    CalculatedVal = 0.0

    for Key in CaliberationParameters:

        if str(CaliberationParameters[Key]).find("-") != -1:
            TempElementList = str(CaliberationParameters[Key]).split("+")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters[str(element)]

            for index in TempElementList:
                CalculatedVal -= float(index)
            CaliberationParameters[Key] = CalculatedVal

    CalculatedVal = 1.0

    for Key in CaliberationParameters:
        if str(CaliberationParameters[Key]).find("*") != -1:

            TempElementList = str(CaliberationParameters[Key]).split("*")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    # print(TempElementList.index(element))
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters[str(element)]

            for element in TempElementList:
                CalculatedVal *= float(element)
            CaliberationParameters[Key] = CalculatedVal

    CalculatedVal = 0.0

    for Key in CaliberationParameters:
        if str(CaliberationParameters[Key]).find("/") != -1:
            TempElementList = str(CaliberationParameters[Key]).split("/")

            for element in TempElementList:
                if re.search('[a-zA-Z]', str(element)) is not None:
                    # print(TempElementList.index(element))
                    TempElementList[int(TempElementList.index(element))] = CaliberationParameters[str(element)]
            indexDiv = 0
            while indexDiv in range(len(TempElementList)) and int(indexDiv) < len(TempElementList) - 1:
                CalculatedVal = (float(TempElementList[indexDiv]) / float(TempElementList[indexDiv + 1]))
                indexDiv += 1
            CaliberationParameters[Key] = CalculatedVal
    CalculatedVal = 0.0

    #print(CaliberationParameters)

    for Key in CaliberationParameters:
        if re.search('[a-zA-Z]', str(CaliberationParameters[Key])) is not None:
            # print(re.search('[a-zA-Z]', str(CaliberationParameters[Key])))
            try:
                CaliberationParameters[Key] = CaliberationParameters[str(CaliberationParameters[Key])];
            except:
                ParseButtonConfig("red")
                #print(Key)
                ErrorPrompt("Parsing Error", "MSGID482 Error parsing Parameter values in Provided Parameter Sheet")

    for Key in CaliberationParameters:
        if str(CaliberationParameters[Key]).find(",") != -1:
            try:
                CaliberationParameters[Key] = str(CaliberationParameters[Key]).replace("[", "{").replace("]", "}");
            except:
                ParseButtonConfig("red")
                ErrorPrompt("Parsing Error", "MSGID490 Error parsing Parameter values in Provided Parameter Sheet")

    # print(CaliberationParameters)
    ParamsheetParseFlag = 1
    ParseButtonConfig("green")

GUITopFrame = tkinter.Tk()  # Creates Object for Top level Tool GUI
GUITopFrame.config(bg="white")
GUITopFrame.title('CalAppend')
GUITopFrame.resizable(False, False)  # x,y resizabling disabled
GUITopFrame.geometry('517x430')

labelframe1 = LabelFrame(GUITopFrame, text="Caliberations Update", bg="white",
                         fg="black")
labelframe1.grid(row=0, columnspan=5, sticky='W', \
                 padx=5, pady=5, ipadx=5, ipady=5)

labelframe2 = LabelFrame(GUITopFrame, text="Caliberations Comparison Report", bg="white",
                         fg="black",width=100)
labelframe2.grid(row=1, columnspan=5, sticky='W', \
                 padx=5, pady=5, ipadx=5, ipady=5)


Label(labelframe1, text="1. Select Caliberation Sheet ", bg="white").grid(row=0,
                                                                          columnspan=1,
                                                                          sticky='W', \
                                                                          padx=15,
                                                                          pady=5,
                                                                          ipadx=5,
                                                                          ipady=2)
Label(labelframe1, text="3. Select Model Generated Code files", bg="white").grid(row=2,
                                                                                 column=0,
                                                                                 columnspan=1,
                                                                                 sticky='W', \
                                                                                 padx=15,
                                                                                 pady=5,
                                                                                 ipadx=5,
                                                                                 ipady=2)
Label(labelframe1, text="2. Parse Caliberations ", bg="white").grid(row=1,
                                                                    column=0,
                                                                    columnspan=1,
                                                                    sticky='W', \
                                                                    padx=15,
                                                                    pady=5,
                                                                    ipadx=5,
                                                                    ipady=2)

CurrentCALUpdate = Label(GUITopFrame, text=" ", bg="white",fg = "green")
CurrentCALUpdate.grid(row=3,column=0,
                                                                    columnspan=5,
                                                                    sticky='W',)

ParamSheetButton = Button(labelframe1, text="Browse ...", fg="white", bg="steel blue",
                        relief="raised", command=ParameterSheetSelect)
ParamSheetButton.grid(row=0, column=1, columnspan=1, sticky='W', \
                                                                            padx=5, pady=5, ipadx=5, ipady=2)
CodeButton = Button(labelframe1, text="Browse ...", fg="white", bg="steel blue",
                        relief="raised", command=UpdateCaliberationsCodeFiles)
CodeButton.grid(row=2, column=1, columnspan=1,sticky='W',padx=5, pady=5, ipadx=5, ipady=2)

RunButtonUpdate = Button(labelframe1, text="Run          ", fg="white", bg="steel blue",
                      relief="raised", command=HandleCodefile)
RunButtonUpdate.grid(row=3, column=1, columnspan=1,sticky='W',padx=5, pady=5, ipadx=5, ipady=2)
ParseButton = Button(labelframe1, text="Parse       ", fg="white", bg="steel blue",
                      relief="raised", command=ParseNCalculateCaliberations)
ParseButton.grid(row=1, column=1, columnspan=1,sticky='W',padx=5, pady=5, ipadx=5, ipady=2)

ProgressStatus = ttk.Progressbar(GUITopFrame, style='text.Horizontal.TProgressbar', length=172, \
                                 cursor='exchange', mode="determinate", orient=tkinter.HORIZONTAL)
ProgressStatus.grid(row=2, column=0,columnspan=2,sticky='W',padx=5,ipadx=5)

style = ttk.Style(GUITopFrame)
style.layout('text.Horizontal.TProgressbar',
             [('Horizontal.Progressbar.trough',
               {'children': [('Horizontal.Progressbar.pbar',
                              {'side': 'left', 'sticky': 'ns'})],
                'sticky': 'nswe'}),
              ('Horizontal.Progressbar.label', {'sticky': ''})])
UpdateProgressStatus(100, 0)

#ProgressStatusComparison.setVisibility(View.GONE);

TuneSelected0 = ttk.Combobox(labelframe1, cursor='arrow', values=TuneColumns, width=25, state="disabled",
                             justify="left")
TuneSelected0.grid(column=2, row=0)
TuneSelected0.current(0)


try:
    GUITopFrame.iconbitmap("Logo_CalAppend.ico")
except:
    ErrorPrompt("File Missing", "CalAppend Tool Logo Logo_CalAppend.ico missing from current \
    directory.\nGUI will be rendered with Incomplete Graphics")

try:
    logo = PhotoImage(file="Logo_KPIT.png")
    logolabel = Label(GUITopFrame, image=logo, bg="white", fg="white", anchor="se").grid(row=2,
                                                                                         column=3,sticky='E',ipadx=90)
except:
    ErrorPrompt("File Missing", "Organisation Logo Logo_KPIT.png missing from current \
	directory.\nGUI will be rendered with Incomplete Graphics")



#############################################################################################################
Label(labelframe2, text="1. Parameter Sheet 1", bg="white").grid(row=0,
                                                                    column=0,
                                                                    columnspan=1,
                                                                    sticky='W', \
                                                                    padx=15,
                                                                    pady=5,
                                                                    ipadx=5,
                                                                    ipady=2)
Label(labelframe2, text="2. Parameter Sheet 2", bg="white").grid(row=1,
                                                                    column=0,
                                                                    columnspan=1,
                                                                    sticky='W', \
                                                                    padx=15,
                                                                    pady=5,
                                                                    ipadx=5,
                                                                    ipady=2)

BrowseCalSheet1 = Button(labelframe2, text="Browse...", fg="white", bg="steel blue",
                      relief="raised",command = lambda:DiferenceParamsheetRead(1))
BrowseCalSheet1.grid(row=0, column=2, columnspan=1,sticky='W',padx=5, pady=5, ipadx=5, ipady=2)

BrowseCalSheet2 = Button(labelframe2, text="Browse...", fg="white", bg="steel blue",
                      relief="raised",command = lambda:DiferenceParamsheetRead(2))
BrowseCalSheet2.grid(row=1, column=2, columnspan=1,sticky='W',padx=5, pady=1, ipadx=5, ipady=2)

GenerateDiffReport = Button(labelframe2, text="Generate", fg="white", bg="steel blue",
                      relief="raised" ,command = GenerateComparisonReport)
GenerateDiffReport.grid(row=2, column=2, columnspan=2,sticky='W',padx=5, pady=1, ipadx=5, ipady=2)

def clear_widget1(event):
    # will clear out any entry boxes defined below when the user shifts
    # focus to the widgets defined below
    if EntryCalSheet1Var == labelframe2.focus_get() and EntryCalSheet1Var.get() == 'Version ':
        EntryCalSheet1Var.delete(0, END)

def clear_widget2(event):
    # will clear out any entry boxes defined below when the user shifts
    # focus to the widgets defined below
    if EntryCalSheet1Var == labelframe2.focus_get() and EntryCalSheet2Var.get() == 'Version ':
        EntryCalSheet2Var.delete(0, END)

def repopulate_defaults1(event):
    # will repopulate the default text previously inside the entry boxes defined below if
    # the user does not put anything in while focused and changes focus to another widget
    if EntryCalSheet1Var != labelframe2.focus_get() and EntryCalSheet1Var.get() == '':
        EntryCalSheet1Var.insert(0, 'Version ')

EntryCalSheet1Var = Entry(labelframe2,fg="white", bg="grey",width="10")
EntryCalSheet1Var.grid(row=0, column=1, columnspan=1,sticky='W',padx=5, pady=1, ipadx=5, ipady=2)
EntryCalSheet1Var.insert(0, 'Version ')
EntryCalSheet2Var = Entry(labelframe2,fg="white", bg="grey",width="10")
EntryCalSheet2Var.grid(row=1, column=1, columnspan=1,sticky='W',padx=5, pady=1, ipadx=5, ipady=2)
EntryCalSheet2Var.insert(0, 'Version ')

TuneSelected1 = ttk.Combobox(labelframe2, cursor='arrow', values=TuneColumns, width=26, state="disabled",
                             justify="left")
TuneSelected1.grid(column=3, row=0)
TuneSelected1.current(0)

TuneSelected2 = ttk.Combobox(labelframe2, cursor='arrow', values=TuneColumns, width=26, state="disabled",
                             justify="left")
TuneSelected2.grid(column=3, row=1)
TuneSelected2.current(0)


GUITopFrame.mainloop()


